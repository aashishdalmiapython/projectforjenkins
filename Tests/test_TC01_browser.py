from selenium.webdriver import Firefox
import pytest
import openpyxl
from Utilities import Screenshot


@pytest.fixture(scope="module")
def setup():
    global driver
    global objwbook
    global wsheet
    global wait
    #wait = WebDriverWait(driver,30)
    objwbook = openpyxl.Workbook()
    wsheet = objwbook.create_sheet("TC01")
    driver = Firefox(executable_path="./Drivers/geckodriver.exe")
    driver.get("https://www.thetestingworld.com/testings/")
    driver.maximize_window()
    driver.implicitly_wait(30)
    yield
    driver.quit()
    objwbook.save("./Utilities/openpyxlwrite.xlsx")


@pytest.mark.High
def test_Registration_tab_heading(setup):
    name = "Test Registration tab heading"
    text = driver.find_element_by_xpath("//label[@for='tab1']").text
    assert text == "Register"
    wsheet.cell(1,1).value = "Test Registration tab heading is Pass"
    Screenshot.take_screenshot(driver, name)

@pytest.mark.High
def test_Registration_username_flabel(setup):
    name = "Test Registration Username Field Label"
    flabel = driver.find_element_by_name("fld_username").get_attribute("placeholder")
    assert flabel == "myusername"
    wsheet.cell(1, 2).value = name + "Pass"
    Screenshot.take_screenshot(driver, name)
