from selenium.webdriver import Firefox
import pytest
import openpyxl

@pytest.fixture(scope="module")
def setup():
    global driver
    global objwbook
    global wsheet
    objwbook = openpyxl.Workbook()
    wsheet = objwbook.create_sheet("TC02")
    driver = Firefox(executable_path="./Drivers/geckodriver.exe")
    driver.get("https://www.thetestingworld.com/testings/")
    driver.maximize_window()
    driver.implicitly_wait(30)
    yield
    driver.quit()
    objwbook.save("./Utilities/openpyxlwrite.xlsx")


@pytest.mark.Regression
def test_login_tab_heading(setup):
    text = driver.find_element_by_xpath("//label[@for='tab1']").text
    assert text == "Register"
    wsheet.cell(1,1).value = "test_Registration_tab_heading is Pass"

@pytest.mark.Smoke
def test_login_username_flabel(setup):
    flabel = driver.find_element_by_name("fld_username").get_attribute("placeholder")
    assert flabel == "myusername"
    wsheet.cell(1, 2).value = "test_Registration_username_flabel is Pass"